import openpyxl
import pandas as pd
import numpy as np
import camelot

#methods for export pdf, extract info from tables by using camelot library and import to excel file
def export_pdf(path_pdf):
    data = camelot.read_pdf(path_pdf, pages='all')
    return data


def import_to_excel(dframe: pd.DataFrame, path: str):
    dframe.to_excel(path,
                    sheet_name='data',
                    columns=['Кредит',
                             'Дата операции',
                             'Дебет',
                             'Дата операции',
                             'Назначение платежа'],
                    header=False,
                    index=False,
                    startrow=2,
                    startcol=3
                    )

#method for result file processing
def excel_processing(bank_name: str,
                     address: str,
                     account: str,
                     path: str):
    account_type = 'р/с ' + account
    workbook = openpyxl.load_workbook(path)
    result_sheet = workbook['data']
    last_row = len(list(result_sheet.rows))
    for row in range(last_row)[2:]:
        row = row + 1
        result_sheet.cell(row=row, column=1, value=bank_name)
        result_sheet.cell(row=row, column=2, value=address)
        result_sheet.cell(row=row, column=3, value=account_type)

        if result_sheet.cell(row=row,column=4).value is None:
            result_sheet.cell(row=row, column=5, value='')

    result_sheet.cell(row=1, column=1, value='Наименование банка (кредитной организации)')
    result_sheet.cell(row=1, column=2, value='Местонахождение')
    result_sheet.cell(row=1, column=3, value='Вид и реквизиты счета')
    result_sheet.cell(row=1, column=4, value='Приход, руб')
    result_sheet.cell(row=1, column=5, value='Дата поступления')
    result_sheet.cell(row=1, column=6, value='Расход, руб')
    result_sheet.cell(row=1, column=7, value='Дата платежа')
    result_sheet.cell(row=1, column=8, value='Обоснование')
    result_sheet.cell(row=2, column=1, value=1)
    result_sheet.cell(row=2, column=2, value=2)
    result_sheet.cell(row=2, column=3, value=3)
    result_sheet.cell(row=2, column=4, value=4)
    result_sheet.cell(row=2, column=5, value=5)
    result_sheet.cell(row=2, column=6, value=6)
    result_sheet.cell(row=2, column=7, value=7)
    result_sheet.cell(row=2, column=8, value=8)
    result_sheet.column_dimensions['A'].width = 22
    result_sheet.column_dimensions['B'].width = 35
    result_sheet.column_dimensions['C'].width = 23
    result_sheet.column_dimensions['E'].width = 10
    result_sheet.column_dimensions['G'].width = 10
    result_sheet.column_dimensions['H'].width = 13

    workbook.save(path)

#class processing extracted data - merge data to dataframe, clearning information from extra columns and rows, renaming columns
#class of processing format depend on bank name, because registers from different banks there are differnt formats
class Cv_processing:
    def __init__(self, data):
        self.data=data

    def uralsib_bank(self) -> pd.DataFrame:
        df = pd.DataFrame
        df_result = pd.DataFrame
        df_r = pd.DataFrame
        df_r = self.data[0].df

        df_result = df_r.rename(columns={df_r.columns[0]: 'Номер документа',
                                         df_r.columns[1]: 'Дата документа',
                                         df_r.columns[2]: 'Дата операции',
                                         df_r.columns[3]: 'Наименование',
                                         df_r.columns[4]: 'Счет',
                                         df_r.columns[5]: 'ИНН котрагента',
                                         df_r.columns[6]: 'Банк контрагента',
                                         df_r.columns[7]: 'Дебет',
                                         df_r.columns[8]: 'Кредит',
                                         df_r.columns[9]: 'Курс ЦБ на дату операции',
                                         df_r.columns[10]: 'Назначение платежа'})

        for tbl in self.data[1:]:
            df = tbl.df
            df_rename = df.rename(columns={df.columns[0]: 'Номер документа',
                                           df.columns[1]: 'Дата документа',
                                           df.columns[2]: 'Дата операции',
                                           df.columns[3]: 'Наименование',
                                           df.columns[4]: 'Счет',
                                           df.columns[5]: 'ИНН котрагента',
                                           df.columns[6]: 'Банк контрагента',
                                           df.columns[7]: 'Дебет',
                                           df.columns[8]: 'Кредит',
                                           df.columns[9]: 'Курс ЦБ на дату операции',
                                           df.columns[10]: 'Назначение платежа'})

            df_result = pd.concat([df_result, df_rename], ignore_index=True)

        df_result['Назначение платежа'].replace('', np.nan, inplace=True)
        df_result['Назначение платежа'].replace('11', np.nan, inplace=True)
        df_result.dropna(subset='Назначение платежа', inplace=True)

        df_result['Банк контрагента'].replace('', np.nan, inplace=True)
        df_result.dropna(subset='Банк контрагента', inplace=True)

        return df_result

    def alpha_bank(self) -> pd.DataFrame:
        df = pd.DataFrame
        df_result = pd.DataFrame
        df_r = pd.DataFrame
        df_r = self.data[0].df

        df_result = df_r.rename(columns={df_r.columns[0]: 'Дата операции',
                                         df_r.columns[1]: 'Номер документа',
                                         df_r.columns[2]: 'Дебет',
                                         df_r.columns[3]: 'Кредит',
                                         df_r.columns[4]: 'ИНН контрагента, счет, наименование',
                                         df_r.columns[5]: 'Банк контрагента',
                                         df_r.columns[6]: 'Назначение платежа',
                                         df_r.columns[7]: 'Код дебитора',
                                         df_r.columns[8]: 'Тип документа'})

        for tbl in self.data[1:]:
            df = tbl.df
            df_rename = df.rename(columns={df.columns[0]: 'Дата операции',
                                           df.columns[1]: 'Номер документа',
                                           df.columns[2]: 'Дебет',
                                           df.columns[3]: 'Кредит',
                                           df.columns[4]: 'ИНН контрагента, счет, наименование',
                                           df.columns[5]: 'Банк контрагента',
                                           df.columns[6]: 'Назначение платежа',
                                           df.columns[7]: 'Код дебитора',
                                           df.columns[8]: 'Тип документа'})

            df_result = pd.concat([df_result, df_rename], ignore_index=True)
        df_result['Банк контрагента'].replace('', np.nan, inplace=True)
        df_result.dropna(subset='Банк контрагента', inplace=True)

        df_result['Дата операции'].replace('', np.nan, inplace=True)
        df_result.dropna(subset='Дата операции', inplace=True)

        return df_result



